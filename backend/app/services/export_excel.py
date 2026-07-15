from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook

from app.models import Invoice
from app.schemas import AccountingEntry


def invoice_to_excel(invoice: Invoice) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction"
    rows = [
        ("Fournisseur", invoice.supplier),
        ("Date", invoice.invoice_date),
        ("Numéro", invoice.invoice_number),
        ("HT", invoice.amount_ht),
        ("TVA", invoice.amount_tva),
        ("TTC", invoice.amount_ttc),
        ("Taux TVA", invoice.vat_rate),
        ("Type", invoice.document_type),
        ("Confiance", invoice.confidence_score),
        ("Statut", invoice.status),
    ]
    ws.append(["Champ", "Valeur"])
    for row in rows:
        ws.append(list(row))

    ws2 = wb.create_sheet("Ecriture")
    ws2.append(["Compte", "Libellé", "Débit", "Crédit"])
    if invoice.accounting_entry:
        entry = AccountingEntry.model_validate_json(invoice.accounting_entry)
        for line in entry.lines:
            ws2.append([line.account, line.label, line.debit, line.credit])

    ws3 = wb.create_sheet("Anomalies")
    ws3.append(["Anomalies"])
    for a in json.loads(invoice.anomalies or "[]"):
        ws3.append([a])
    ws3.append(["Champs manquants"])
    for m in json.loads(invoice.missing_fields or "[]"):
        ws3.append([m])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def invoices_to_excel(invoices: list[Invoice]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"
    ws.append(
        [
            "ID",
            "Fichier",
            "Fournisseur",
            "Date",
            "Numéro",
            "HT",
            "TVA",
            "TTC",
            "Taux",
            "Type",
            "Confiance",
            "Statut",
            "À vérifier",
        ]
    )
    for inv in invoices:
        ws.append(
            [
                inv.id,
                inv.filename,
                inv.supplier,
                inv.invoice_date,
                inv.invoice_number,
                inv.amount_ht,
                inv.amount_tva,
                inv.amount_ttc,
                inv.vat_rate,
                inv.document_type,
                inv.confidence_score,
                inv.status,
                "Oui" if inv.needs_review else "Non",
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()